from openpyxl import load_workbook


def get_sheet_data(file_path, sheet_name):
    workbook = load_workbook(filename=file_path, data_only=True)

    if sheet_name not in workbook.sheetnames:
        return None

    sheet = workbook[sheet_name]
    sheet_data = []

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: row[i] for i in range(len(headers))}
        sheet_data.append(row_data)

    return sheet_data

from PythonInit.GetRelativePathForProjectFiles import get_data_path
if __name__ == "__main__":
    file_path = get_data_path("test-data.xlsx", "resources")
    sheet_name = 'req-payload'
    data = get_sheet_data(file_path, sheet_name)
    print(data)
