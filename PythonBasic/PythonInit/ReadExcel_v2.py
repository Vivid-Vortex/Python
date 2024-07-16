import re
from collections import defaultdict

from openpyxl import load_workbook


def get_sheet_data(file_path, sheet_name):
    workbook = load_workbook(filename=file_path, data_only=True)

    if sheet_name not in workbook.sheetnames:
        print(f"Sheet {sheet_name} not found in workbook.")
        return None

    sheet = workbook[sheet_name]
    sheet_data = []

    # Identify headers and index columns
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index_columns = {i: headers[i] for i in range(len(headers)) if headers[i] and len(headers[i]) == 1 and headers[i].islower()}

    print(f"Headers: {headers}")
    print(f"Index columns: {index_columns}")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}

        # Replace placeholders in headers with actual values
        processed_row_data = {}
        for header, value in row_data.items():
            if header not in index_columns.values():
                new_header = header
                for col_idx, col_name in index_columns.items():
                    if col_name in row_data and row_data[col_name] is not None:
                        new_header = re.sub(rf'\[{col_name}\]', f'[{row_data[col_name]}]', new_header)
                processed_row_data[new_header] = value

        if processed_row_data:  # Ensure we only add non-empty dictionaries
            sheet_data.append(processed_row_data)

    return sheet_data

from collections import defaultdict

def merge_dicts_by_id(data):
    merged_data = defaultdict(dict)

    for entry in data:
        id_value = entry.get('id')
        if id_value is not None:
            for key, value in entry.items():
                if value is not None:
                    merged_data[id_value][key] = value
            if 'id' not in merged_data[id_value]:
                merged_data[id_value]['id'] = id_value

    return list(merged_data.values())


def remove_none_values(data):
    cleaned_data = []
    for entry in data:
        cleaned_entry = {key: value for key, value in entry.items() if value is not None}
        cleaned_data.append(cleaned_entry)
    return cleaned_data

def process_dicts(data):
  """
  Processes a list of dictionaries, filtering out keys with "request" prefix,
  moving them to a separate dictionary within each main dictionary,
  and returns the modified list.

  Args:
    data: A list of dictionaries.

  Returns:
    A new list of dictionaries with the processed data.
  """

  processed_data = []
  for entry in data:
    request_data = {}
    filtered_data = {}
    for key, value in entry.items():
      if key.startswith("request"):
        request_data[key] = value
      else:
        filtered_data[key] = value
    filtered_data["requestJsonPaths"] = request_data  # Add "request" dictionary
    processed_data.append(filtered_data)
  return processed_data

from PythonInit.GetRelativePathForProjectFiles import get_data_path
if __name__ == "__main__":
    file_path = get_data_path("test-data-v2.xlsx", "resources")
    sheet_name = 'req-payload2'
    data = get_sheet_data(file_path, sheet_name)
    data = merge_dicts_by_id(data)
    # data = remove_none_values(data)
    data = process_dicts(data)
    print(data)
